from __future__ import annotations

import csv
import io
import re
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from langgraph.graph import END, START, StateGraph
from openpyxl import load_workbook
from typing_extensions import TypedDict

from .models import (
    ColumnMapping,
    ColumnMappingDecision,
    ImportPreview,
    ImportedCase,
    SheetReport,
    TableLocationDecision,
    STANDARD_FIELDS,
)
from .llm import create_deepseek_model, deepseek_enabled


REGION_HEADER_WINDOW = 25  # header candidates are scored within the first rows of a region
HEADER_MIN_SCORE = 1.0  # a header row must score at least this much
DATA_OVERLAP_MIN = 0.5  # a following row must fill this fraction of the header columns to count as data
LLM_TABLE_MAX = 3  # at most this many tables per sheet are accepted from the model
LLM_PREVIEW_ROWS = 60  # rows sent to the model for table location
LLM_TABLE_CONFIDENCE_CAP = 0.7  # mappings on model-located tables are never trusted above this
LLM_TABLE_HEADER_PROXIMITY = 3  # a model table within this many rows of a rule table is a duplicate


ALIASES = {
    "case_id": {"case id", "caseid", "test case id", "testcase id", "用例id", "用例编号", "编号"},
    "case_type": {"case type", "type", "platform", "channel", "用例类型", "类型", "端"},
    "description": {"description", "title", "case name", "test case", "scenario", "scenario name", "summary", "用例描述", "描述", "用例名称", "场景"},
    "preconditions": {"pre conditions", "preconditions", "pre condition", "prerequisite", "prerequisites", "前置条件", "执行前置条件"},
    "test_steps": {"test steps", "steps", "step", "actions", "procedure", "测试步骤", "操作步骤", "步骤"},
    "test_data": {"test data", "data", "data set", "dataset", "data point", "测试数据", "数据集", "测试资料"},
    "expected_result": {"expected result", "expected results", "expected", "outcome", "assertion", "预期结果", "期望结果"},
    "priority": {"priority", "severity", "importance", "优先级", "重要级别"},
}


def normalize(value: Any) -> str:
    text = str(value or "").strip().lower().replace("_", "-")
    return re.sub(r"[\s\-–—/:()]+", " ", text).strip()


def rule_mapping(headers: list[str]) -> list[ColumnMapping]:
    mappings: list[ColumnMapping] = []
    used: set[str] = set()
    for header in headers:
        normalized = normalize(header)
        exact = next((field for field, aliases in ALIASES.items() if normalized in aliases), None)
        if exact and exact not in used:
            mappings.append(ColumnMapping(source_column=header, target_field=exact, confidence=1, reason="Known field alias"))
            used.add(exact)
            continue
        partial = next(
            (
                field
                for field, aliases in ALIASES.items()
                if field not in used and any(len(alias) >= 4 and (alias in normalized or normalized in alias) for alias in aliases)
            ),
            None,
        )
        if partial:
            mappings.append(ColumnMapping(source_column=header, target_field=partial, confidence=.82, reason="Similar field name"))
            used.add(partial)
        else:
            mappings.append(ColumnMapping(source_column=header, target_field=None, confidence=1, reason="Preserved as an extra field"))
    return mappings


def semantic_mapping(headers: list[str], samples: list[dict[str, Any]], current: list[ColumnMapping]) -> list[ColumnMapping]:
    if not deepseek_enabled():
        return current
    ambiguous = [item.source_column for item in current if item.target_field is None]
    if not ambiguous:
        return current
    try:
        from langchain.agents import create_agent
        agent = create_agent(
            model=create_deepseek_model(),
            tools=[],
            response_format=ColumnMappingDecision,
            system_prompt=(
                "Map spreadsheet columns to the QA Orbit schema. Allowed targets: "
                + ", ".join(STANDARD_FIELDS)
                + ". Use null for unrelated fields. Never invent a source column or map two columns to one target."
            ),
        )
        result = agent.invoke({"messages": [{"role": "user", "content": f"Headers: {headers}\nSample rows: {samples[:3]}"}]})
        decision = result["structured_response"]
        proposed = {item.source_column: item for item in decision.mappings if item.source_column in ambiguous}
        used = {item.target_field for item in current if item.target_field}
        merged: list[ColumnMapping] = []
        for item in current:
            candidate = proposed.get(item.source_column)
            if candidate and candidate.target_field in STANDARD_FIELDS and candidate.target_field not in used:
                candidate.confidence = min(candidate.confidence, .79)
                candidate.reason = "LangChain semantic mapping"
                merged.append(candidate)
                used.add(candidate.target_field)
            else:
                merged.append(item)
        return merged
    except Exception:
        return current


def read_workbook(filename: str, content: bytes) -> list[tuple[str, list[list[Any]]]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig", errors="replace")
        return [(Path(filename).stem, list(csv.reader(io.StringIO(text))))]
    if suffix == ".xls":
        import xlrd

        book = xlrd.open_workbook(file_contents=content)
        return [(sheet.name, [sheet.row_values(index) for index in range(sheet.nrows)]) for sheet in book.sheets()]
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("Supported files are .xlsx, .xls, .xlsm and .csv")
    book = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    sheets: list[tuple[str, list[list[Any]]]] = []
    for sheet in book.worksheets:
        rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
        for merged in list(sheet.merged_cells.ranges):
            value = rows[merged.min_row - 1][merged.min_col - 1]
            for row_index in range(merged.min_row - 1, merged.max_row):
                for column_index in range(merged.min_col - 1, merged.max_col):
                    rows[row_index][column_index] = value
        sheets.append((sheet.title, rows))
    return sheets


def header_score(row: list[Any]) -> float:
    values = [normalize(value) for value in row if str(value or "").strip()]
    if len(values) < 2:
        return 0
    known = sum(any(value in aliases or any(alias in value for alias in aliases if len(alias) >= 4) for aliases in ALIASES.values()) for value in values)
    return known * 4 + min(len(values), 8) * .25


def is_active_row(row: list[Any]) -> bool:
    """A row that is dense enough to belong to a table (>= 2 non-empty cells)."""
    return sum(1 for value in row if str(value or "").strip()) >= 2


def table_regions(rows: list[list[Any]]) -> list[tuple[int, int]]:
    """Split a sheet into table-like regions.

    Rows with >= 2 non-empty cells are members; a single sparse row between two
    members is bridged, while 2+ consecutive sparse rows split regions. This keeps
    one table with an occasional blank row intact while separating independent
    tables that may sit anywhere in the sheet (title blocks, notes, etc.).
    """
    member = [is_active_row(row) for row in rows]
    for index in range(len(member)):
        if not member[index] and index > 0 and index + 1 < len(member) and member[index - 1] and member[index + 1]:
            member[index] = True
    regions: list[tuple[int, int]] = []
    start: int | None = None
    for index, active in enumerate(member + [False]):
        if active and start is None:
            start = index
        elif not active and start is not None:
            regions.append((start, index - 1))
            start = None
    return regions


@dataclass
class TableResult:
    """One detected table inside a sheet, using absolute row indices."""

    header_index: int
    headers: list[str]
    data_rows: list[dict[str, Any]]  # records keyed by header, each carrying __source_row__
    located_by: str = "rules"  # "rules" | "llm"
    reason: str = ""


def analyze_region(rows: list[list[Any]], start: int, end: int) -> TableResult | None:
    """Detect a single table inside rows[start..end] (absolute indices)."""
    window = min(end + 1, start + REGION_HEADER_WINDOW)
    candidates = [(index, header_score(rows[index])) for index in range(start, window)]
    header_index, score = max(candidates, key=lambda item: item[1])
    if score < HEADER_MIN_SCORE:
        return None
    raw_headers = rows[header_index]
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(raw_headers):
        header = str(value or f"Column {index + 1}").strip()
        seen[header] = seen.get(header, 0) + 1
        headers.append(f"{header} ({seen[header]})" if seen[header] > 1 else header)
    header_columns = [index for index, value in enumerate(raw_headers) if str(value or "").strip()]
    data_rows: list[dict[str, Any]] = []
    aligned = False
    for source_index in range(header_index + 1, end + 1):
        row = rows[source_index]
        if not is_active_row(row):
            continue
        if not aligned:
            filled = sum(1 for column in header_columns if column < len(row) and str(row[column] or "").strip())
            if header_columns and filled / len(header_columns) < DATA_OVERLAP_MIN:
                continue  # first dense row below the header does not line up; keep scanning
            aligned = True
        values = list(row) + [None] * max(0, len(headers) - len(row))
        record = {headers[index]: values[index] for index in range(len(headers)) if values[index] not in (None, "")}
        if record:
            record["__source_row__"] = source_index + 1  # 1-based, matches the Excel row shown to users
            data_rows.append(record)
    if not aligned:
        has_following = any(is_active_row(row) for row in rows[header_index + 1 : end + 1])
        reason = "Header row found but rows below it do not align with the columns" if has_following else "Header row found but no data rows below it"
        return TableResult(header_index=header_index, headers=headers, data_rows=data_rows, reason=reason)
    return TableResult(header_index=header_index, headers=headers, data_rows=data_rows)


def detect_tables(rows: list[list[Any]]) -> list[TableResult]:
    """Find every rule-detected table in a sheet (rules only, no LLM)."""
    tables: list[TableResult] = []
    for start, end in table_regions(rows):
        table = analyze_region(rows, start, end)
        if table is not None:
            tables.append(table)
    return tables


def sheet_has_content(rows: list[list[Any]]) -> bool:
    return any(is_active_row(row) for row in rows)


def _build_preview_payload(name: str, rows: list[list[Any]]) -> dict[str, Any]:
    preview: list[dict[str, Any]] = []
    for index, row in enumerate(rows[:LLM_PREVIEW_ROWS], start=1):
        cells = [[column + 1, str(value)[:40]] for column, value in enumerate(row) if str(value or "").strip()]
        if len(cells) >= 2:
            preview.append({"row": index, "cells": cells})
    return {"sheet": name, "rows": preview}


def locate_tables_with_llm(name: str, rows: list[list[Any]]) -> list[TableResult]:
    """Ask the model to locate test-case tables when rule detection found none.

    Only a compact preview of the first rows (non-empty cells with coordinates)
    is sent. Located blocks are re-validated by the standard region analysis, so a
    hallucinated table without real aligned data rows is dropped.
    """
    if not deepseek_enabled():
        return []
    try:
        from langchain.agents import create_agent

        agent = create_agent(
            model=create_deepseek_model(),
            tools=[],
            response_format=TableLocationDecision,
            system_prompt=(
                "You locate test-case tables inside spreadsheet sheets. A test-case table is a rectangular block "
                "with a header row whose columns describe test cases (id, description, steps, expected result, "
                "priority, ...) and data rows directly beneath it. Ignore titles, instructions, and unrelated "
                "content. Return each table with 1-based row numbers relative to the whole sheet and the exact "
                "header labels as listed. Use an empty tables array when there is none. Never invent row numbers "
                "outside the provided rows."
            ),
        )
        result = agent.invoke({"messages": [{"role": "user", "content": f"Workbook sheet preview:\n{_build_preview_payload(name, rows)}"}]})
        decision = result["structured_response"]
    except Exception:
        return []
    tables: list[TableResult] = []
    for located in decision.tables[:LLM_TABLE_MAX]:
        header_index = located.header_row - 1
        first_data = located.first_data_row - 1
        last_data = located.last_data_row - 1
        if not (0 <= header_index < len(rows) and first_data > header_index and last_data >= first_data and last_data < len(rows)):
            continue
        region = analyze_region(rows, header_index, last_data)
        if region is None or not region.data_rows:
            continue
        region.header_index = header_index
        region.located_by = "llm"
        region.reason = f"Table located by the model at row {located.header_row}; verify the header row"
        tables.append(region)
    return tables


def normalize_type(value: Any) -> str:
    text = str(value or "Web").strip().lower()
    if "api" in text or "接口" in text:
        return "API"
    if "mobile" in text or "app" in text or "移动" in text:
        return "Mobile"
    return "Web"


def normalize_priority(value: Any) -> str:
    text = str(value or "P1").strip().upper()
    aliases = {"HIGH": "P0", "CRITICAL": "P0", "最高": "P0", "MEDIUM": "P1", "中": "P1", "LOW": "P2", "低": "P2"}
    return text if text in {"P0", "P1", "P2"} else aliases.get(text, "P1")


class ImportState(TypedDict, total=False):
    filename: str
    content: bytes
    sheets: list[tuple[str, list[list[Any]]]]
    preview: ImportPreview


def parse_node(state: ImportState) -> dict[str, Any]:
    return {"sheets": read_workbook(state["filename"], state["content"])}


def map_node(state: ImportState) -> dict[str, Any]:
    import_id = str(uuid.uuid4())
    all_cases: list[ImportedCase] = []
    reports: list[SheetReport] = []
    warnings: list[str] = []
    llm_assisted: list[str] = []
    next_generated = 1
    for sheet_name, rows in state["sheets"]:
        tables = detect_tables(rows)
        # Rules found no table with data: let the model locate irregular tables
        # that the heuristics could not see (mid-sheet blocks, odd layouts).
        if not any(table.data_rows for table in tables) and sheet_has_content(rows) and deepseek_enabled():
            located = locate_tables_with_llm(sheet_name, rows)
            tables = tables + [
                table
                for table in located
                if all(abs(table.header_index - existing.header_index) > LLM_TABLE_HEADER_PROXIMITY for existing in tables)
            ]
        if not tables:
            reports.append(SheetReport(name=sheet_name, status="skipped", reason="No tabular test case data detected"))
            continue
        for table_index, table in enumerate(tables, start=1):
            mappings = semantic_mapping(table.headers, table.data_rows, rule_mapping(table.headers))
            if table.located_by == "llm":
                mappings = [item.model_copy(update={"confidence": min(item.confidence, LLM_TABLE_CONFIDENCE_CAP)}) for item in mappings]
            target_by_source = {item.source_column: item.target_field for item in mappings}
            confidence_by_source = {item.source_column: item.confidence for item in mappings}
            sheet_count = 0
            for record in table.data_rows:
                standard: dict[str, Any] = {}
                extras: dict[str, Any] = {}
                provenance: dict[str, str] = {}
                confidences: list[float] = []
                for source, value in record.items():
                    if source == "__source_row__":
                        continue
                    target = target_by_source.get(source)
                    if target:
                        standard[target] = value
                        provenance[target] = f"{sheet_name}!{source} row {record['__source_row__']}"
                        confidences.append(confidence_by_source[source])
                    else:
                        extras[source] = value
                if not any(str(standard.get(field, "")).strip() for field in ("description", "test_steps", "expected_result")):
                    continue
                raw_id = str(standard.get("case_id") or "").strip()
                case_id = raw_id or f"IMP-{next_generated:04d}"
                case_warnings = []
                if not raw_id:
                    case_warnings.append("Case ID was generated")
                    next_generated += 1
                if not standard.get("description"):
                    case_warnings.append("Description is missing")
                imported = ImportedCase(
                    case_id=case_id,
                    case_type=normalize_type(standard.get("case_type")),
                    description=str(standard.get("description") or "").strip(),
                    preconditions=str(standard.get("preconditions") or "").strip(),
                    test_steps=str(standard.get("test_steps") or "").strip(),
                    test_data=str(standard.get("test_data") or "").strip(),
                    expected_result=str(standard.get("expected_result") or "").strip(),
                    priority=normalize_priority(standard.get("priority")),
                    extra_fields=extras,
                    source_file=state["filename"],
                    source_sheet=sheet_name,
                    source_row=record["__source_row__"],
                    import_order=len(all_cases) + 1,
                    field_provenance=provenance,
                    mapping_confidence=sum(confidences) / len(confidences) if confidences else .5,
                    warnings=case_warnings,
                )
                all_cases.append(imported)
                sheet_count += 1
            if sheet_count:
                status = "imported"
                reason = ""
            else:
                status = "no-data"
                reason = table.reason or "Header row found but no test cases were produced"
            if table.located_by == "llm" and sheet_count:
                llm_assisted.append(sheet_name)
            reports.append(
                SheetReport(
                    name=sheet_name,
                    status=status,
                    reason=reason,
                    header_row=table.header_index + 1,
                    row_count=sheet_count,
                    mappings=mappings,
                    table_index=table_index,
                )
            )
    if not all_cases:
        warnings.append("No test cases were detected. Check the header row and field names.")
    for name in dict.fromkeys(llm_assisted):
        warnings.append(f"{name}: a table was located with AI assistance — verify the header row and field mapping.")
    table_count = sum(1 for report in reports if report.status in {"imported", "no-data"})
    explanation = [
        f"Scanned {len(state['sheets'])} sheet(s) for table-like regions and recognized {table_count} table(s).",
        "Matched known aliases first, then used the LangChain semantic mapper for ambiguous columns when a model key was available.",
        "Preserved every unmatched column in extra_fields and recorded the source sheet and row for each case.",
    ]
    return {"preview": ImportPreview(import_id=import_id, filename=state["filename"], cases=all_cases, sheets=reports, warnings=warnings, explanation=explanation)}


workflow = StateGraph(ImportState)
workflow.add_node("read_workbook", parse_node)
workflow.add_node("map_and_validate", map_node)
workflow.add_edge(START, "read_workbook")
workflow.add_edge("read_workbook", "map_and_validate")
workflow.add_edge("map_and_validate", END)
IMPORT_GRAPH = workflow.compile()


def build_preview(filename: str, content: bytes) -> ImportPreview:
    return IMPORT_GRAPH.invoke({"filename": filename, "content": content})["preview"]
