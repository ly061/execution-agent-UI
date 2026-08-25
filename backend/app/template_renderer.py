from __future__ import annotations

from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .models import GeneratedCase


FIELD_VALUES = {
    "case_id": lambda case: case.case_id,
    "case_type": lambda case: case.case_type,
    "description": lambda case: case.title,
    "preconditions": lambda case: case.preconditions,
    "test_steps": lambda case: case.test_steps,
    "test_data": lambda case: "",
    "expected_result": lambda case: case.expected_result,
    "priority": lambda case: case.priority,
}


def _copy_row_style(sheet, source_row: int, target_row: int, width: int) -> None:
    if source_row < 1 or source_row > sheet.max_row:
        return
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, width + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def render_cases(profile: dict[str, Any], cases: list[GeneratedCase]) -> tuple[str, bytes]:
    artifact = Path(str(profile.get("artifact_path") or ""))
    if artifact.suffix.lower() == ".xlsx" and artifact.is_file():
        workbook = load_workbook(artifact)
    else:
        workbook = Workbook()
        default = workbook.active
        default.title = str((profile.get("sheets") or [{}])[0].get("name") or "Test Cases")

    sheet_profile = next(
        (item for item in profile.get("sheets") or [] if item.get("field_mapping")),
        None,
    ) or {
        "name": workbook.sheetnames[0],
        "header_row": 1,
        "columns": ["Case ID", "Description", "Type", "Priority", "Preconditions", "Test Steps", "Expected Result"],
        "field_mapping": {
            "Case ID": "case_id",
            "Description": "description",
            "Type": "case_type",
            "Priority": "priority",
            "Preconditions": "preconditions",
            "Test Steps": "test_steps",
            "Expected Result": "expected_result",
        },
    }
    name = str(sheet_profile.get("name") or workbook.sheetnames[0])
    sheet = workbook[name] if name in workbook.sheetnames else workbook.create_sheet(name)
    header_row = int(sheet_profile.get("header_row") or 1)
    columns = list(sheet_profile.get("columns") or [])
    mapping = dict(sheet_profile.get("field_mapping") or {})
    if not columns:
        columns = list(mapping)
    for column_index, column_name in enumerate(columns, start=1):
        sheet.cell(header_row, column_index).value = column_name

    first_data_row = header_row + 1
    style_source_row = first_data_row if sheet.max_row >= first_data_row else header_row
    existing_last_row = sheet.max_row
    for case_index, case in enumerate(cases):
        row = first_data_row + case_index
        if row > existing_last_row:
            _copy_row_style(sheet, style_source_row, row, len(columns))
        for column_index, column_name in enumerate(columns, start=1):
            field = mapping.get(column_name)
            value = FIELD_VALUES[field](case) if field in FIELD_VALUES else ""
            sheet.cell(row, column_index).value = value
    for row in range(first_data_row + len(cases), existing_last_row + 1):
        for column in range(1, len(columns) + 1):
            sheet.cell(row, column).value = None

    output = BytesIO()
    workbook.save(output)
    requested = str(profile.get("filename") or "generated-test-cases.xlsx")
    stem = Path(requested).stem or "generated-test-cases"
    return f"{stem}-generated.xlsx", output.getvalue()
