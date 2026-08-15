from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook

import app.importer as importer
from app import database
from app.main import app


def preview_xlsx(book: Workbook, tmp_path, monkeypatch) -> dict:
    monkeypatch.delenv("DEEPSEEK_API_KEY", raising=False)
    monkeypatch.setattr(database, "DATA_DIR", tmp_path)
    monkeypatch.setattr(database, "DB_PATH", tmp_path / "test.sqlite")
    database.init_db()
    buffer = BytesIO()
    book.save(buffer)
    with TestClient(app) as client:
        response = client.post(
            "/api/imports/preview",
            files={"file": ("cases.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200, response.text
        return response.json()


def test_header_only_sheet_reports_no_data_with_schema(tmp_path, monkeypatch):
    book = Workbook()
    schema = book.active
    schema.title = "Schema Only"
    schema.append(["Test Case ID", "Scenario Name", "Platform", "Actions", "Outcome"])
    cases = book.create_sheet("Cases")
    cases.append(["Case ID", "Description", "Steps", "Expected Result", "Priority"])
    cases.append(["C-1", "Login", "Open site", "Dashboard opens", "P1"])

    preview = preview_xlsx(book, tmp_path, monkeypatch)

    assert len(preview["cases"]) == 1
    schema_report = next(sheet for sheet in preview["sheets"] if sheet["name"] == "Schema Only")
    assert schema_report["status"] == "no-data"
    assert schema_report["header_row"] == 1
    assert schema_report["row_count"] == 0
    assert schema_report["mappings"], "schema-only sheet must still report its column mappings"
    assert any(mapping["source_column"] == "Scenario Name" and mapping["target_field"] == "description" for mapping in schema_report["mappings"])
    cases_report = next(sheet for sheet in preview["sheets"] if sheet["name"] == "Cases")
    assert cases_report["status"] == "imported"
    assert cases_report["row_count"] == 1


def test_table_below_first_25_rows_is_found(tmp_path, monkeypatch):
    book = Workbook()
    ws = book.active
    ws.title = "Late Table"
    for index in range(30):
        ws.append([f"junk {index}"])
    ws.append(["Case ID", "Description", "Steps", "Expected Result"])
    ws.append(["L-1", "Deep table", "Do the thing", "OK"])

    preview = preview_xlsx(book, tmp_path, monkeypatch)

    report = preview["sheets"][0]
    assert report["status"] == "imported"
    assert report["header_row"] == 31
    assert preview["cases"][0]["case_id"] == "L-1"
    assert preview["cases"][0]["source_row"] == 32


def test_two_tables_in_one_sheet_each_imported(tmp_path, monkeypatch):
    book = Workbook()
    ws = book.active
    ws.title = "Two Tables"
    ws.append(["Case ID", "Description", "Steps", "Expected Result", "Priority"])
    ws.append(["T1-1", "Login", "Open", "OK", "P1"])
    ws.append(["T1-2", "Logout", "Close", "OK", "P1"])
    ws.append(["spacer"])
    ws.append(["spacer"])
    ws.append(["Case ID", "Description", "Steps", "Expected Result"])
    ws.append(["T2-1", "Register", "Sign up", "OK"])

    preview = preview_xlsx(book, tmp_path, monkeypatch)

    reports = [sheet for sheet in preview["sheets"] if sheet["name"] == "Two Tables"]
    assert len(reports) == 2
    first, second = reports
    assert first["table_index"] == 1 and first["status"] == "imported" and first["row_count"] == 2 and first["header_row"] == 1
    assert second["table_index"] == 2 and second["status"] == "imported" and second["row_count"] == 1 and second["header_row"] == 6
    assert len(preview["cases"]) == 3
    assert {case["case_id"] for case in preview["cases"]} == {"T1-1", "T1-2", "T2-1"}
    assert all(case["source_sheet"] == "Two Tables" for case in preview["cases"])


def test_merged_title_block_does_not_split_the_table(tmp_path, monkeypatch):
    book = Workbook()
    ws = book.active
    ws.title = "Merged Title"
    ws.append(["Release regression cases"])
    ws.merge_cells("A1:E1")
    ws.append(["Test Case ID", "Scenario Name", "Platform", "Actions", "Outcome"])
    ws.append(["WEB-1", "Login", "Web", "Open login", "Dashboard opens"])

    preview = preview_xlsx(book, tmp_path, monkeypatch)

    report = preview["sheets"][0]
    assert report["status"] == "imported"
    assert report["header_row"] == 2
    assert preview["cases"][0]["case_id"] == "WEB-1"


def test_llm_located_table_is_imported_with_capped_confidence(tmp_path, monkeypatch):
    book = Workbook()
    ws = book.active
    ws.title = "Messy"
    for index in range(10):
        ws.append(["note", f"line {index}"])  # dense rows but nothing scores as a header -> rules find nothing
    # no real header/data rows: the model must supply the table boundaries

    def fake_locate(name, rows):
        assert name == "Messy"
        return [
            importer.TableResult(
                header_index=10,
                headers=["Case ID", "Description", "Steps", "Expected Result", "Priority"],
                data_rows=[{"Case ID": "M-1", "Description": "AI found me", "Steps": "Run it", "Expected Result": "OK", "Priority": "P1", "__source_row__": 11}],
                located_by="llm",
            )
        ]

    monkeypatch.setattr(importer, "deepseek_enabled", lambda: True)
    monkeypatch.setattr(importer, "locate_tables_with_llm", fake_locate)
    preview = preview_xlsx(book, tmp_path, monkeypatch)

    assert len(preview["cases"]) == 1
    assert preview["cases"][0]["case_id"] == "M-1"
    assert preview["cases"][0]["mapping_confidence"] <= importer.LLM_TABLE_CONFIDENCE_CAP
    report = next(sheet for sheet in preview["sheets"] if sheet["name"] == "Messy")
    assert report["status"] == "imported"
    assert report["header_row"] == 11
    assert any("Messy" in warning for warning in preview["warnings"])


def test_rules_still_work_when_llm_key_absent(tmp_path, monkeypatch):
    book = Workbook()
    ws = book.active
    ws.title = "Plain"
    ws.append(["Case ID", "Description", "Steps", "Expected Result"])
    ws.append(["P-1", "Simple", "Go", "OK"])

    preview = preview_xlsx(book, tmp_path, monkeypatch)

    assert len(preview["cases"]) == 1
    assert preview["sheets"][0]["status"] == "imported"
