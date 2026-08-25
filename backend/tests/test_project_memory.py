import json
import asyncio
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from app import database, deep_orchestrator, generation_agent, project_memory
from app.main import app
from app.models import ProjectAgentRequest


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "登录用例"
    sheet.append(["用例编号", "用例描述", "前置条件", "测试步骤", "预期结果", "优先级", "模块"])
    sheet.append(
        [
            "LOGIN-001",
            "登录-有效账号-进入首页",
            "用户已注册",
            "1. 打开登录页\n2. 输入有效账号\n3. 点击登录",
            "用户登录成功并进入首页",
            "P0",
            "认证",
        ]
    )
    sheet["A2"].font = Font(bold=True)
    sheet.append(
        [
            "LOGIN-002",
            "登录-错误密码-提示失败",
            "用户已注册",
            "1. 打开登录页\n2. 输入错误密码\n3. 点击登录",
            "页面提示账号或密码错误",
            "P1",
            "认证",
        ]
    )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_learns_project_template_and_style_as_candidate(tmp_path, monkeypatch):
    monkeypatch.setattr(database, "DATA_DIR", tmp_path)
    monkeypatch.setattr(database, "DB_PATH", tmp_path / "memory.sqlite")
    database.init_db()

    with TestClient(app) as client:
        response = client.post(
            "/api/projects/project-a/learning/cases",
            data={"user_id": "qa-user"},
            files={
                "file": (
                    "approved-cases.xlsx",
                    _workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200, response.text
        learned = response.json()
        assert learned["project_id"] == "project-a"
        assert learned["imported_count"] == 2
        assert learned["style_profile"]["language"] == "zh-CN"
        assert learned["style_profile"]["step_style"] == "numbered_multiline"
        assert learned["template_profile"]["sheets"][0]["columns"][0] == "用例编号"
        assert "模块" in learned["template_profile"]["extra_fields"]
        assert "artifact_path" not in learned["template_profile"]
        assert learned["memory_candidates"][0]["status"] == "candidate"

        profiles = client.get("/api/projects/project-a/profiles", params={"user_id": "qa-user"}).json()
        assert profiles["style_profile"]["sample_count"] == 2
        assert profiles["template_profile"]["filename"] == "approved-cases.xlsx"
        assert "artifact_path" not in profiles["template_profile"]

        exported = client.post(
            "/api/projects/project-a/export",
            json={
                "filename": "login-regression",
                "cases": [
                    {
                        "case_id": "GEN-001",
                        "title": "登录-锁定账号-提示失败",
                        "case_type": "Web",
                        "priority": "P0",
                        "preconditions": "账号已锁定",
                        "test_steps": "1. 输入锁定账号\n2. 点击登录",
                        "expected_result": "页面提示账号已锁定",
                        "requirement": "账号锁定",
                    }
                ],
            },
        )
        assert exported.status_code == 200, exported.text
        assert "login-regression.xlsx" in exported.headers["content-disposition"]
        rendered = load_workbook(BytesIO(exported.content))
        result_sheet = rendered["登录用例"]
        assert result_sheet["A1"].value == "用例编号"
        assert result_sheet["A2"].value == "GEN-001"
        assert result_sheet["B2"].value == "登录-锁定账号-提示失败"
        assert result_sheet["A2"].font.bold is True
        assert result_sheet["A3"].value is None


def test_generation_retrieval_is_strictly_project_scoped(tmp_path, monkeypatch):
    monkeypatch.setattr(database, "DATA_DIR", tmp_path)
    monkeypatch.setattr(database, "DB_PATH", tmp_path / "scopes.sqlite")
    database.init_db()
    database.save_memory(
        project_id="project-a",
        memory_type="semantic",
        content="Accounts lock after five failed login attempts.",
        status="active",
    )
    database.save_memory(
        project_id="project-b",
        memory_type="semantic",
        content="Project B uses a completely different biometric login rule.",
        status="active",
    )
    captured = {}

    async def fake_call_model(api_key, messages):
        captured["payload"] = json.loads(messages[-1]["content"])
        return generation_agent._parse_decision(
            {
                "action": "ask",
                "message": "Choose a platform.",
                "questions": [{"id": "q1", "question": "Web or Mobile?", "options": ["Web", "Mobile"]}],
            }
        )

    monkeypatch.setattr(generation_agent, "_call_model", fake_call_model)
    with TestClient(app) as client:
        response = client.post(
            "/api/generation/sessions",
            data={"text": "Test failed login attempts.", "project_id": "project-a", "user_id": "qa-user"},
            headers={"X-DeepSeek-API-Key": "test-key"},
        )
        assert response.status_code == 200, response.text
        context = captured["payload"]["project_context"]
        contents = [item["content"] for item in context["memories"]]
        assert any("five failed" in content for content in contents)
        assert all("Project B" not in content for content in contents)
        row = database.get_generation_session(response.json()["session_id"])
        assert row["project_id"] == "project-a"
        assert row["user_id"] == "qa-user"


def test_memory_requires_activation_before_retrieval(tmp_path, monkeypatch):
    monkeypatch.setattr(database, "DATA_DIR", tmp_path)
    monkeypatch.setattr(database, "DB_PATH", tmp_path / "approval.sqlite")
    database.init_db()
    candidate = database.save_memory(
        project_id="project-a",
        memory_type="procedural",
        content="Always write expected results per step.",
        status="candidate",
    )
    assert project_memory.retrieve_context("project-a", None, "write cases")["memories"] == []

    with TestClient(app) as client:
        activated = client.patch(
            f"/api/projects/project-a/memories/{candidate['id']}",
            json={"status": "active"},
        )
        assert activated.status_code == 200
    memories = project_memory.retrieve_context("project-a", None, "write cases")["memories"]
    assert memories[0]["content"] == "Always write expected results per step."


def test_quality_review_suggests_missing_negative_and_boundary_cases():
    decision = generation_agent._parse_decision(
        {
            "action": "generate",
            "cases": [
                {
                    "title": "User creates a password",
                    "case_type": "Web",
                    "priority": "P1",
                    "test_steps": "1. Enter a valid password\n2. Submit",
                    "expected_result": "Password is saved",
                    "requirement": "Password length is at least 8 characters",
                }
            ],
        }
    )
    reviewed = generation_agent._with_quality_review(
        decision,
        "Password length must be at least 8 characters.",
        {"style_profile": None},
    )
    titles = {suggestion.title for suggestion in reviewed.suggestions}
    assert "Add negative-path coverage" in titles
    assert "Add boundary-value cases" in titles
    assert "Teach this project its preferred style" in titles


def test_structured_patch_preserves_untouched_cases_and_stable_ids():
    first = generation_agent.GeneratedCase(
        case_id="case-first",
        title="Valid login",
        case_type="Web",
        priority="P0",
        test_steps="1. Log in",
        expected_result="Dashboard opens",
    )
    second = generation_agent.GeneratedCase(
        case_id="case-second",
        title="Invalid login",
        case_type="Web",
        priority="P1",
        test_steps="1. Use a wrong password",
        expected_result="Login is rejected",
    )
    decision = generation_agent._parse_decision(
        {
            "action": "update",
            "message": "Raised the invalid-login priority.",
            "operations": [
                {"op": "update", "case_id": "case-second", "field": "priority", "value": "P0"}
            ],
        }
    )
    updated = generation_agent.apply_case_operations([first, second], decision.operations)
    assert updated[0].model_dump() == first.model_dump()
    assert updated[1].case_id == "case-second"
    assert updated[1].priority == "P0"
    assert updated[1].title == second.title


def test_deep_agent_supervisor_compiles_with_project_scoped_tools():
    supervisor = deep_orchestrator.build_project_supervisor("project-a", "qa-user", "test-key")
    assert supervisor.name == "qa-orbit-supervisor"
    graph = supervisor.get_graph()
    assert graph.nodes, "Deep Agents must compile to a runnable LangGraph"


def test_deep_agent_rejects_generation_session_from_another_project(tmp_path, monkeypatch):
    monkeypatch.setattr(database, "DATA_DIR", tmp_path)
    monkeypatch.setattr(database, "DB_PATH", tmp_path / "deep-agent-scope.sqlite")
    database.init_db()
    database.create_generation_session(
        "session-b",
        "requirements.txt",
        "Private requirements for project B",
        {"status": "generated", "action": "generate", "message": "done", "cases": []},
        project_id="project-b",
    )
    request = ProjectAgentRequest(
        message="Review this session",
        generation_session_id="session-b",
    )
    try:
        asyncio.run(deep_orchestrator.advise_project("project-a", request, "test-key"))
    except ValueError as error:
        assert "does not belong" in str(error)
    else:
        raise AssertionError("Cross-project generation sessions must be rejected before the model is called")
