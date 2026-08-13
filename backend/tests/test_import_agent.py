from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app import database
from app.main import app


def workbook_bytes() -> bytes:
    book = Workbook()
    notes = book.active
    notes.title = "Instructions"
    notes.append(["QA workbook", "Complete all required fields"])
    web = book.create_sheet("Web Cases")
    web.append(["Release regression cases"])
    web.merge_cells("A1:C1")
    web.append(["Test Case ID", "Scenario Name", "Platform", "Prerequisite", "Actions", "Outcome", "Severity", "User Name", "Owner"])
    for index in range(1, 8):
        web.append([f"WEB-{index}", f"Login scenario {index}", "Web", "User exists", f"1. Open login\n2. Submit user {index}", "Dashboard opens", "High" if index == 1 else "Medium", f"User {index}", "QA"])
    api = book.create_sheet("接口用例")
    api.append(["用例编号", "用例描述", "类型", "测试步骤", "预期结果", "优先级", "备注"])
    api.append(["API-1", "Validate profile endpoint", "API", "GET /profile", "Returns 200", "P1", "Keep this"])
    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def test_multisheet_preview_confirm_chat_and_undo(tmp_path, monkeypatch):
    monkeypatch.delenv("DEEPSEEK_API_KEY", raising=False)
    monkeypatch.setattr(database, "DATA_DIR", tmp_path)
    monkeypatch.setattr(database, "DB_PATH", tmp_path / "test.sqlite")
    database.init_db()
    with TestClient(app) as client:
        preview_response = client.post(
            "/api/imports/preview",
            files={"file": ("cases.xlsx", workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert preview_response.status_code == 200, preview_response.text
        preview = preview_response.json()
        assert len(preview["cases"]) == 8
        assert preview["cases"][6]["extra_fields"]["User Name"] == "User 7"
        assert any(sheet["name"] == "Instructions" and sheet["status"] == "skipped" for sheet in preview["sheets"])
        web_sheet = next(sheet for sheet in preview["sheets"] if sheet["name"] == "Web Cases")
        assert any(mapping["source_column"] == "Scenario Name" and mapping["target_field"] == "description" for mapping in web_sheet["mappings"])

        corrected = client.post(f"/api/imports/{preview['import_id']}/chat", json={"message": "第七条 case 导入的不太对，user name 应该用 Lisa"})
        assert corrected.status_code == 200, corrected.text
        assert corrected.json()["changes"][0]["after"] == "Lisa"
        assert corrected.json()["cases"][0]["extra_fields"]["User Name"] == "Lisa"

        confirmed = client.post(f"/api/imports/{preview['import_id']}/confirm")
        assert confirmed.status_code == 200
        assert confirmed.json()["imported_count"] == 8
        seventh = next(case for case in confirmed.json()["cases"] if case["import_order"] == 7)
        assert seventh["extra_fields"]["User Name"] == "Lisa"

        undone = client.post(f"/api/imports/{preview['import_id']}/chat", json={"message": "撤销"})
        assert undone.status_code == 200
        assert undone.json()["changes"][0]["after"] == "User 7"
